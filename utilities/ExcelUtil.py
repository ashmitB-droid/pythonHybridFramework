import openpyxl


def get_row_count(path, sheet_name):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet_name]
    return sheet.max_row


def get_col_count(path, sheet_name):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet_name]
    return sheet.max_row


def get_cell_data(path, sheet_name, row_number, column_number):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet_name]
    return sheet.cell(row=row_number,col=column_number).value


def get_cell_data(path, sheet_name, row_number, column_number, data):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet_name]
    sheet.cell(row=row_number,col=column_number).value=data
    workbook.save(path)


def get_data_from_excel(path, sheet_name):
    result_list = []
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet_name]
    total_rows = sheet.max_row
    total_cols = sheet.max_column
    for row_num in range(2,total_rows+1):
        row_list = []
        for col_num in range(1,total_cols+1):
            row_list.append(sheet.cell(row=row_num,column=col_num).value)
        result_list.append(row_list)
    return result_list






